import asyncio
import datetime
import math
import re
import subprocess
import os
import random
import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
import threading
import queue
import logging

from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# Configuration
BATCH_SIZE = 20
WAIT_BETWEEN_BATCHES = 30  # En secondes
WAIT_BETWEEN_PLATES_MIN = 8  # Délai minimum en secondes
WAIT_BETWEEN_PLATES_MAX = 10  # Délai maximum en secondes

# Créer une file d'attente pour communiquer entre asyncio et tkinter
progress_queue = queue.Queue()

# Configurer le module logging
logger = logging.getLogger("PlateChecker")
logger.setLevel(logging.INFO)
# Créer un gestionnaire qui envoie les logs à la file d'attente
class QueueHandler(logging.Handler):
    def __init__(self, queue):
        super().__init__()
        self.queue = queue

    def emit(self, record):
        self.queue.put(("log", self.format(record)))

# Ajouter le gestionnaire à logger
queue_handler = QueueHandler(progress_queue)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
queue_handler.setFormatter(formatter)
logger.addHandler(queue_handler)

def load_plates_from_excel(filename="BaseDePlaques.xlsx"):
    """
    Lit un fichier Excel contenant 3 colonnes:
      - IMMATRICULATION
      - Categorie vehicule
      - Proprietaire

    Retourne une liste de tuples (immatriculation, categorie, proprietaire).
    """
    wb = load_workbook(filename)
    ws = wb.active

    data = []
    # On part de row=2 pour ignorer la ligne d'en-tête
    for row in ws.iter_rows(min_row=2, values_only=True):
        # row sera un tuple du type (IMMATRICULATION, Categorie vehicule, Proprietaire)
        if not row or all(cell is None for cell in row):
            continue  # ignorer les lignes vides
        immat, categorie, proprietaire = row
        # On stocke ce trio pour usage ultérieur
        data.append((str(immat).strip(), str(categorie).strip(), str(proprietaire).strip()))
    return data

async def check_plate(browser, immat, categorie, proprietaire, progress_queue):
    """
    Lance la vérification pour 1 plaque.
    Retourne (immat, categorie, proprietaire, status, montant, date_passage).
    """
    page = await browser.new_page()
    status = "Erreur"
    montant = "N/A"
    date_passage = "N/A"  # Renamed field: "Date de passage"

    try:
        # 1) Goto
        await page.goto("https://www.sanef.com/client/index.html#basket",
                        wait_until="networkidle",
                        timeout=20000)
        logger.info(f"{immat}: Accès à la page SANEF.")

        # 2) Cookies
        try:
            await page.click(".tarteaucitronCTAButton", timeout=8000)
            # Delay after clicking accept cookies
            await asyncio.sleep(random.uniform(1, 2))
            logger.info(f"{immat}: Acceptation des cookies.")
        except:
            logger.warning(f"{immat}: Bouton cookies non trouvé ou déjà accepté.")

        # Envoyer le véhicule actuel en cours de traitement
        progress_queue.put(("current", immat))

        # 3) Fill the plate
        await page.wait_for_selector("input.input-no-focus", timeout=12000)
        await page.fill("input.input-no-focus", immat)
        # Delay after filling the plate
        await asyncio.sleep(random.uniform(1, 2))
        logger.info(f"{immat}: Saisie de la plaque.")

        # 4) Click + short pause
        await page.click("text=Vérifier mes péages à payer")
        # Delay after clicking the verify button
        await asyncio.sleep(random.uniform(1, 2))
        await page.wait_for_timeout(4000)
        logger.info(f"{immat}: Clic sur 'Vérifier mes péages à payer'.")

        # 5) Body text
        body_text = await page.text_content("body") or ""

        # 6) Determine status
        if "Aucun passage en attente de paiement" in body_text:
            status = "Rien à payer"
            montant = "0 €"
            logger.info(f"{immat} ({categorie}, {proprietaire}): Rien à payer")
        else:
            logger.info(f"{immat} ({categorie}, {proprietaire}): Péages dus")

            # Try to find total amount by selector
            total_elem = await page.query_selector("span.total-amount")
            if total_elem:
                extracted = (await total_elem.text_content() or "").strip()
                montant = extracted
                logger.info(f"{immat}: Montant trouvé via span.total-amount: {montant}")
            else:
                # Fallback Regex for amounts
                pattern = r"\b(\d{1,3},\d{2}\s?€)\b"
                matches = re.findall(pattern, body_text)
                if matches:
                    # Convert all matches to float for comparison
                    amounts = []
                    for m in matches:
                        # Nettoyer la chaîne pour convertir en float
                        clean_m = m.replace('€', '').replace(',', '.').strip()
                        try:
                            amount = float(clean_m)
                            amounts.append(amount)
                        except ValueError:
                            pass  # Ignorer les valeurs qui ne peuvent pas être converties

                    if amounts:
                        # Sélectionner le montant le plus élevé
                        max_amount = max(amounts)
                        montant = f"{max_amount:.2f} €"
                        logger.info(f"{immat}: Montant total trouvé: {montant}")
                    else:
                        montant = "Inconnu"
                        logger.warning(f"{immat}: Impossible de trouver un montant valide.")
                else:
                    montant = "Inconnu"
                    logger.warning(f"{immat}: Impossible de trouver le montant.")

            # 7) Try to parse all “date de passage” from the page
            #    Example regex for a French format DD/MM/YYYY
            date_passage_pattern = r"(\d{2}/\d{2}/\d{4})"
            matches_date = re.findall(date_passage_pattern, body_text)
            if matches_date:
                # Enlever les doublons et trier les dates (optionnel)
                unique_dates = sorted(set(matches_date), key=lambda date: datetime.datetime.strptime(date, "%d/%m/%Y"))
                date_passage = ", ".join(unique_dates)
                num_peages = len(unique_dates)
                status = f"Péages dus: {num_peages}"
                logger.info(f"{immat}: Dates de passage trouvées : {date_passage}")
                logger.info(f"{immat}: Nombre de péages dus : {num_peages}")
            else:
                date_passage = "N/A"
                num_peages = 0
                status = "Péages dus: 0"
                logger.warning(f"{immat}: Impossible de trouver la date de passage.")

    except Exception as e:
        logger.error(f"{immat}: Erreur - {e}")
    finally:
        await page.close()
        logger.info(f"{immat}: Fermeture de la page.")

    # Wait between each plate with random delay to simulate human behavior
    await asyncio.sleep(random.uniform(WAIT_BETWEEN_PLATES_MIN, WAIT_BETWEEN_PLATES_MAX))  # Sleep between 8 and 10 seconds

    # Mettre à jour la barre de progression en ajoutant 1
    progress_queue.put(("progress", 1))

    # Return full info
    return (immat, categorie, proprietaire, status, montant, date_passage)


async def get_browser_instance(p):
    """
    Tente Chrome, puis Edge, puis Firefox.
    """
    try:
        logger.info("Tentative de lancement de Chrome...")
        return await p.chromium.launch(channel="chrome", headless=False)
    except Exception as e:
        logger.error(f"Échec du lancement de Chrome: {e}")
        pass
    try:
        logger.info("Tentative de lancement de Edge...")
        return await p.chromium.launch(channel="msedge", headless=False)
    except Exception as e:
        logger.error(f"Échec du lancement de Edge: {e}")
        pass
    try:
        logger.info("Tentative de lancement de Firefox...")
        return await p.firefox.launch(headless=False)
    except Exception as e:
        logger.error(f"Échec du lancement de Firefox: {e}")
        pass
    raise RuntimeError("Aucun navigateur supporté trouvé.")

async def process_batch_sequential(browser, plates_data, progress_queue):
    """
    'plates_data' est une liste de tuples (immat, categorie, proprietaire).
    On les traite un par un, en séquentiel.
    """
    results = []
    for (immat, cat, prop) in plates_data:
        result = await check_plate(browser, immat, cat, prop, progress_queue)
        results.append(result)
    return results

async def main_async(plates_data, progress_queue):
    ############################################################################
    # 3) Normal script logic (Playwright checks)
    ############################################################################
    async with async_playwright() as p:
        all_results = []
        total_plates = len(plates_data)
        num_batches = math.ceil(total_plates / BATCH_SIZE)

        for i in range(num_batches):
            # Extract the portion for this batch
            start = i * BATCH_SIZE
            end = start + BATCH_SIZE
            batch_plates = plates_data[start:end]
            logger.info(f"=== Batch {i+1}/{num_batches}, {len(batch_plates)} plaques ===")
            # Process them sequentially
            try:
                browser = await get_browser_instance(p)
            except RuntimeError as e:
                logger.error(f"Erreur lors du lancement du navigateur: {e}")
                return

            batch_results = await process_batch_sequential(browser, batch_plates, progress_queue)
            all_results.extend(batch_results)

            await browser.close()
            logger.info(f"=== Batch {i+1}/{num_batches} terminé. Navigateur fermé. ===")

            # Wait between batches with random delay
            if i < num_batches - 1:
                wait_time = random.uniform(WAIT_BETWEEN_BATCHES - 5, WAIT_BETWEEN_BATCHES + 5)  # Exemple: 25 à 35 secondes
                logger.info(f"Attente de {wait_time:.2f} secondes avant le prochain batch...")
                await asyncio.sleep(wait_time)

            # Ajouter une pause de 5 minutes après chaque 5 batches (100 véhicules)
            if (i + 1) % 5 == 0 and (i + 1) < num_batches:
                logger.info("=== 100 véhicules traités. Attente de 5 minutes avant de continuer ===")
                await asyncio.sleep(300)  # 5 minutes en secondes

    # all_results est maintenant une liste de tuples :
    # (immat, categorie, proprietaire, status, montant, date_passage)

    # Sort by immatriculation (index=0)
    results_sorted = sorted(all_results, key=lambda x: x[0])
    logger.info("\n=== Récapitulatif final ===")
    for immat, cat, prop, status, montant, date_passage in results_sorted:
        logger.info(f"{immat} ({cat}, {prop}) -> {status} (Montant: {montant}), Date de passage: {date_passage}")

    ############################################################################
    # 4) Write results to resultats.xlsx with conditional formatting
    ############################################################################
    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats péages"

    ws.column_dimensions["A"].width = 20  # Heure et Date
    ws.column_dimensions["B"].width = 20  # Immatriculation
    ws.column_dimensions["C"].width = 30  # Catégorie
    ws.column_dimensions["D"].width = 20  # Propriétaire
    ws.column_dimensions["E"].width = 30  # Statut
    ws.column_dimensions["F"].width = 15  # Montant
    ws.column_dimensions["G"].width = 30  # Date de passage

    # Define fills for conditional formatting
    peages_due_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light Red
    rien_a_payer_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Light Green

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Header row
    ws.append(["Heure et Date",
               "IMMATRICULATION",
               "Categorie vehicule",
               "Proprietaire",
               "Statut",
               "Montant",
               "Date de passage"])

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Fill rows with conditional formatting
    for immat, cat, prop, status, montant, date_passage in results_sorted:
        ws.append([now_str, immat, cat, prop, status, montant, date_passage])
        current_row = ws.max_row
        statut_cell = ws.cell(row=current_row, column=5)  # Column E: Statut

        if statut_cell.value.startswith("Péages dus"):
            statut_cell.fill = peages_due_fill
        else:
            statut_cell.fill = rien_a_payer_fill

    excel_filename = "resultats.xlsx"
    wb.save(excel_filename)
    logger.info(f"\nFichier Excel enregistré sous {excel_filename}")

    ############################################################################
    # 5) Open resultats.xlsx at the end (Windows only)
    ############################################################################
    try:
        os.startfile(excel_filename)
        logger.info(f"Ouverture de {excel_filename}...")
    except AttributeError:
        logger.warning("os.startfile() n'est pas supporté sur ce système d'exploitation.")
    except Exception as e:
        logger.error(f"Impossible d'ouvrir {excel_filename}: {e}")

    # Indiquer que le traitement est terminé
    progress_queue.put(("done", None))

def run_asyncio_loop(plates_data, progress_queue):
    asyncio.run(main_async(plates_data, progress_queue))

def create_gui(total_plates):
    root = tk.Tk()
    root.title("Avancement du Traitement des Véhicules")
    root.geometry("700x500")
    root.resizable(False, False)

    # Titre
    title_label = tk.Label(root, text="Traitement des Véhicules", font=("Helvetica", 16))
    title_label.pack(pady=10)

    # Barre de progression
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=600, mode="determinate")
    progress_bar.pack(pady=10)
    progress_bar["maximum"] = total_plates

    # Pourcentage
    percent_label = tk.Label(root, text="0%", font=("Helvetica", 12))
    percent_label.pack()

    # Véhicule actuel
    current_vehicle_label = tk.Label(root, text="Véhicule actuel: Aucun", font=("Helvetica", 12))
    current_vehicle_label.pack(pady=5)

    # Bouton Afficher/Masquer les logs
    log_visible = tk.BooleanVar(value=True)
    def toggle_logs():
        if log_visible.get():
            log_text.pack_forget()
            toggle_button.config(text="Afficher les logs")
        else:
            log_text.pack(pady=10)
            toggle_button.config(text="Masquer les logs")
        log_visible.set(not log_visible.get())

    toggle_button = tk.Button(root, text="Masquer les logs", command=toggle_logs)
    toggle_button.pack()

    # Zone de log (ScrolledText)
    log_text = scrolledtext.ScrolledText(root, width=80, height=20, state='disabled', wrap='word')
    log_text.pack(pady=10)

    # Message de fin
    end_message_label = tk.Label(root, text="", font=("Helvetica", 12), fg="blue")
    end_message_label.pack(pady=5)

    def update_progress():
        while not progress_queue.empty():
            message, value = progress_queue.get()
            if message == "progress":
                progress_bar["value"] += value
                percent = (progress_bar["value"] / progress_bar["maximum"]) * 100
                percent_label.config(text=f"{percent:.2f}%")
            elif message == "current":
                current_vehicle_label.config(text=f"Véhicule actuel: {value}")
            elif message == "done":
                current_vehicle_label.config(text="Traitement terminé")
                end_message_label.config(text="Toutes les plaques ont été traitées.")
            elif message == "log":
                log_text.configure(state='normal')
                log_text.insert(tk.END, value + "\n")
                log_text.configure(state='disabled')
                log_text.yview(tk.END)  # Auto-scroll vers le bas
        root.after(100, update_progress)  # Vérifie toutes les 100 ms

    # Démarrer la mise à jour de la barre de progression
    root.after(100, update_progress)

    return root, progress_bar, current_vehicle_label, end_message_label, log_text, toggle_button

def main_gui():
    # Fermer les instances Excel ouvertes
    try:
        subprocess.run(["taskkill", "/IM", "EXCEL.EXE", "/F"], check=True)
        logger.info("Toutes les instances Excel ont été fermées pour libérer 'resultats.xlsx'.")
    except subprocess.CalledProcessError:
        logger.warning("Aucune instance Excel à fermer.")
    except FileNotFoundError:
        logger.warning("Commande 'taskkill' non trouvée.")

    # Charger les plaques depuis Excel
    plates_data = load_plates_from_excel("BaseDePlaques.xlsx")
    if not plates_data:
        logger.error("BaseDePlaques.xlsx est vide ou invalide.")
        return

    total_plates = len(plates_data)

    # Créer la GUI
    root, progress_bar, current_vehicle_label, end_message_label, log_text, toggle_button = create_gui(total_plates)

    # Démarrer l'asyncio loop dans un thread séparé
    thread = threading.Thread(target=run_asyncio_loop, args=(plates_data, progress_queue), daemon=True)
    thread.start()

    # Démarrer la boucle Tkinter
    root.mainloop()

    # Attendre que le thread asyncio se termine
    thread.join()

    # Une fois la boucle Tkinter fermée, le script se termine

if __name__ == "__main__":
    main_gui()
